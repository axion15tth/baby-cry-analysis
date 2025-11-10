from io import BytesIO
from typing import Dict, Any, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.utils.time_utils import seconds_to_absolute_time, format_datetime, format_seconds


class ExcelExporter:
    """Excel形式で解析結果をエクスポート（5シート構成）

    シート構成:
    1. 概要 - ファイル情報と解析サマリー
    2. 泣き声エピソード - 検出されたエピソード一覧
    3. 音響特徴統計 - 各エピソードの統計情報と特殊パラメータ
    4. 音響特徴 - 時系列の音響特徴データ
    5. Cry Units - 各エピソードのCry Unit情報
    """

    def __init__(self):
        self.wb = Workbook()

    def export(
        self,
        result_data: Dict[str, Any],
        file_info: Dict[str, Any],
        recording_start_time: Optional[datetime] = None
    ) -> bytes:
        """
        解析結果をExcel形式でエクスポート

        Args:
            result_data: 解析結果データ
            file_info: ファイル情報
            recording_start_time: 録音開始時刻

        Returns:
            Excelファイルのバイナリデータ
        """
        # デフォルトシートを削除
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # 5シートを作成
        self._create_summary_sheet(result_data, file_info, recording_start_time)
        self._create_episodes_sheet(result_data, recording_start_time)
        self._create_statistics_sheet(result_data)
        self._create_features_sheet(result_data, recording_start_time)
        self._create_cry_units_sheet(result_data, recording_start_time)

        # バイナリデータとして返す
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_summary_sheet(
        self,
        result_data: Dict[str, Any],
        file_info: Dict[str, Any],
        recording_start_time: Optional[datetime]
    ):
        """シート1: 概要"""
        ws = self.wb.create_sheet("概要")

        # タイトル
        ws["A1"] = "赤ちゃん泣き声解析レポート"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:B1")

        # ファイル情報
        row = 3
        ws[f"A{row}"] = "ファイル情報"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        info_data = [
            ("ファイル名", file_info.get("original_filename", "")),
            ("ファイルサイズ", f"{file_info.get('file_size', 0) / 1024:.2f} KB"),
            ("録音開始時刻", format_datetime(recording_start_time) if recording_start_time else "未設定"),
            ("解析日時", file_info.get("analyzed_at", "")),
        ]

        for label, value in info_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # 解析サマリー
        row += 1
        ws[f"A{row}"] = "解析サマリー"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        cry_episodes = result_data.get("cry_episodes", [])
        total_cry_duration = sum(ep["duration"] for ep in cry_episodes)

        summary_data = [
            ("検出エピソード数", len(cry_episodes)),
            ("総泣き時間", f"{total_cry_duration:.2f} 秒"),
            ("平均エピソード長", f"{total_cry_duration / len(cry_episodes):.2f} 秒" if cry_episodes else "0 秒"),
        ]

        for label, value in summary_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # 列幅調整
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40

    def _create_episodes_sheet(
        self,
        result_data: Dict[str, Any],
        recording_start_time: Optional[datetime]
    ):
        """シート2: 泣き声エピソード一覧"""
        ws = self.wb.create_sheet("泣き声エピソード")

        # ヘッダー行
        if recording_start_time:
            headers = [
                "No.", "開始時刻（絶対）", "終了時刻（絶対）",
                "開始時刻（秒）", "終了時刻（秒）", "継続時間（秒）", "信頼度"
            ]
        else:
            headers = [
                "No.", "開始時刻（秒）", "終了時刻（秒）", "継続時間（秒）", "信頼度"
            ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # データ行
        cry_episodes = result_data.get("cry_episodes", [])
        for i, episode in enumerate(cry_episodes, 1):
            start_time = episode["start_time"]
            end_time = episode["end_time"]
            duration = episode["duration"]
            confidence = episode["confidence"]

            row = i + 1
            if recording_start_time:
                abs_start = seconds_to_absolute_time(recording_start_time, start_time)
                abs_end = seconds_to_absolute_time(recording_start_time, end_time)
                ws.cell(row, 1, i)
                ws.cell(row, 2, format_datetime(abs_start))
                ws.cell(row, 3, format_datetime(abs_end))
                ws.cell(row, 4, float(format_seconds(start_time)))
                ws.cell(row, 5, float(format_seconds(end_time)))
                ws.cell(row, 6, float(format_seconds(duration)))
                ws.cell(row, 7, float(f"{confidence:.4f}"))
            else:
                ws.cell(row, 1, i)
                ws.cell(row, 2, float(format_seconds(start_time)))
                ws.cell(row, 3, float(format_seconds(end_time)))
                ws.cell(row, 4, float(format_seconds(duration)))
                ws.cell(row, 5, float(f"{confidence:.4f}"))

        # 列幅自動調整
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_statistics_sheet(
        self,
        result_data: Dict[str, Any]
    ):
        """シート3: 音響特徴統計"""
        ws = self.wb.create_sheet("音響特徴統計")

        # タイトル
        ws["A1"] = "音響特徴統計情報"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")

        row = 3
        statistics = result_data.get("statistics", {})

        for episode_id in sorted(statistics.keys()):
            episode_num = int(episode_id.split("_")[1]) + 1
            episode_stats = statistics[episode_id]

            # エピソードヘッダー
            ws[f"A{row}"] = f"エピソード {episode_num}"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            ws.merge_cells(f"A{row}:F{row}")
            row += 1

            # 音響パラメータの統計量テーブル
            headers = ["パラメータ", "平均", "標準偏差", "最小値", "最大値", "中央値"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            row += 1

            # 音響パラメータのデータ
            acoustic_params = ['f0', 'f1', 'f2', 'f3', 'hnr', 'shimmer', 'jitter', 'intensity']
            param_labels = {
                'f0': 'F0 (Hz)',
                'f1': 'F1 (Hz)',
                'f2': 'F2 (Hz)',
                'f3': 'F3 (Hz)',
                'hnr': 'HNR (dB)',
                'shimmer': 'Shimmer (%)',
                'jitter': 'Jitter (%)',
                'intensity': 'Intensity (dB)'
            }

            for param in acoustic_params:
                if param in episode_stats:
                    stats = episode_stats[param]
                    ws.cell(row, 1, param_labels.get(param, param.upper()))

                    def safe_float(v):
                        return float(f"{v:.4f}") if v is not None else None

                    ws.cell(row, 2, safe_float(stats.get("mean")))
                    ws.cell(row, 3, safe_float(stats.get("std")))
                    ws.cell(row, 4, safe_float(stats.get("min")))
                    ws.cell(row, 5, safe_float(stats.get("max")))
                    ws.cell(row, 6, safe_float(stats.get("median")))
                    row += 1

            # 特殊パラメータセクション
            row += 1
            ws[f"A{row}"] = "特殊パラメータ"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            ws[f"B{row}"] = "値 (%)"
            ws[f"B{row}"].font = Font(bold=True)
            ws[f"B{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            ws.merge_cells(f"C{row}:F{row}")
            row += 1

            special_params = [
                ('high_pitch_pct', 'High-pitch割合'),
                ('hyper_phonation_pct', 'Hyper-phonation割合'),
                ('voiced_pct', '有声音割合'),
                ('unvoiced_pct', '無声音割合')
            ]

            for param_key, param_label in special_params:
                if param_key in episode_stats:
                    ws.cell(row, 1, param_label)
                    value = episode_stats[param_key]
                    ws.cell(row, 2, float(f"{value:.2f}") if value is not None else None)
                    row += 1

            row += 2  # エピソード間にスペース

        # 列幅調整
        ws.column_dimensions["A"].width = 25
        for col in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 15

    def _create_features_sheet(
        self,
        result_data: Dict[str, Any],
        recording_start_time: Optional[datetime]
    ):
        """シート3: 音響特徴詳細（全エピソード）"""
        ws = self.wb.create_sheet("音響特徴")

        # ヘッダー行
        if recording_start_time:
            headers = [
                "エピソードNo.", "時刻（絶対）", "時刻（秒）",
                "F0 (Hz)", "F1 (Hz)", "F2 (Hz)", "F3 (Hz)",
                "HNR (dB)", "Shimmer (%)", "Jitter (%)", "Intensity (dB)"
            ]
        else:
            headers = [
                "エピソードNo.", "時刻（秒）",
                "F0 (Hz)", "F1 (Hz)", "F2 (Hz)", "F3 (Hz)",
                "HNR (dB)", "Shimmer (%)", "Jitter (%)", "Intensity (dB)"
            ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # データ行
        acoustic_features = result_data.get("acoustic_features", {})
        row = 2

        for episode_id, features_list in sorted(acoustic_features.items()):
            # エピソード番号を抽出（"episode_0" -> 1）
            episode_num = int(episode_id.split("_")[1]) + 1

            for feature in features_list:
                time = feature["time"]

                if recording_start_time:
                    abs_time = seconds_to_absolute_time(recording_start_time, time)
                    ws.cell(row, 1, episode_num)
                    ws.cell(row, 2, format_datetime(abs_time))
                    ws.cell(row, 3, float(format_seconds(time)))

                    col_offset = 4
                else:
                    ws.cell(row, 1, episode_num)
                    ws.cell(row, 2, float(format_seconds(time)))
                    col_offset = 3

                # 音響パラメータ
                params = ["f0", "f1", "f2", "f3", "hnr", "shimmer", "jitter", "intensity"]
                for i, param in enumerate(params):
                    value = feature.get(param)
                    if value is not None:
                        ws.cell(row, col_offset + i, float(f"{value:.4f}"))

                row += 1

        # 列幅自動調整
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_cry_units_sheet(
        self,
        result_data: Dict[str, Any],
        recording_start_time: Optional[datetime] = None
    ):
        """Cry Unitsシートを作成"""
        ws = self.wb.create_sheet("Cry Units")

        # ヘッダースタイル
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # ヘッダー
        if recording_start_time:
            headers = [
                "エピソード", "Unit番号", "開始時刻（絶対）", "終了時刻（絶対）",
                "開始時刻（相対秒）", "終了時刻（相対秒）",
                "継続時間（秒）", "有声音/無声音", "平均エネルギー", "ピーク周波数 (Hz)"
            ]
        else:
            headers = [
                "エピソード", "Unit番号", "開始時刻（秒）", "終了時刻（秒）",
                "継続時間（秒）", "有声音/無声音", "平均エネルギー", "ピーク周波数 (Hz)"
            ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # データ行
        cry_units = result_data.get("cry_units", {})
        row = 2

        for episode_id in sorted(cry_units.keys()):
            episode_data = cry_units[episode_id]
            units = episode_data.get("units", [])

            # エピソード番号を抽出（episode_0 -> 1）
            episode_num = int(episode_id.split("_")[1]) + 1

            for i, unit in enumerate(units, 1):
                start_time = unit["start_time"]
                end_time = unit["end_time"]
                duration = unit["duration"]
                is_voiced = unit["is_voiced"]
                mean_energy = unit["mean_energy"]
                peak_frequency = unit["peak_frequency"]

                voicing_label = "有声音" if is_voiced else "無声音"

                ws.cell(row, 1, episode_num)
                ws.cell(row, 2, i)

                if recording_start_time:
                    abs_start = seconds_to_absolute_time(recording_start_time, start_time)
                    abs_end = seconds_to_absolute_time(recording_start_time, end_time)
                    ws.cell(row, 3, format_datetime(abs_start))
                    ws.cell(row, 4, format_datetime(abs_end))
                    ws.cell(row, 5, float(format_seconds(start_time)))
                    ws.cell(row, 6, float(format_seconds(end_time)))
                    ws.cell(row, 7, float(format_seconds(duration)))
                    ws.cell(row, 8, voicing_label)
                    ws.cell(row, 9, float(f"{mean_energy:.6f}"))
                    ws.cell(row, 10, float(f"{peak_frequency:.2f}"))
                else:
                    ws.cell(row, 3, float(format_seconds(start_time)))
                    ws.cell(row, 4, float(format_seconds(end_time)))
                    ws.cell(row, 5, float(format_seconds(duration)))
                    ws.cell(row, 6, voicing_label)
                    ws.cell(row, 7, float(f"{mean_energy:.6f}"))
                    ws.cell(row, 8, float(f"{peak_frequency:.2f}"))

                row += 1

        # 列幅自動調整
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
